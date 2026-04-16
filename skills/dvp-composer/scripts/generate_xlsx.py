#!/usr/bin/env python3
"""DVP Composer - Generate formatted Excel (.xlsx) from JSON spec.

Usage:
    python3 generate_xlsx.py <json_file> [--output <output_path>]

The JSON file must conform to the schema defined in references/excel-spec.md.
"""

import json
import sys
import os
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is not installed. Run: pip3 install openpyxl")
    sys.exit(1)

# ── Style Constants ──────────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SECTION_TITLE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SECTION_TITLE_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79")
LABEL_FONT = Font(name="Calibri", size=11, bold=True)
NORMAL_FONT = Font(name="Calibri", size=11)
ALT_ROW_FILL = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
WRAP_ALIGNMENT = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ── Helpers ──────────────────────────────────────────────────────────────────

def apply_header_style(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def apply_section_title(ws, row, col_count, title):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_count)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = SECTION_TITLE_FONT
    cell.fill = SECTION_TITLE_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = THIN_BORDER


def apply_cell_border(ws, row, col_count):
    for col in range(1, col_count + 1):
        ws.cell(row=row, column=col).border = THIN_BORDER


def auto_width(ws, col_count, max_width=50):
    for col in range(1, col_count + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        adjusted = min(max_len + 4, max_width)
        ws.column_dimensions[get_column_letter(col)].width = max(adjusted, 12)


def apply_alt_row(ws, row, col_count):
    if row % 2 == 0:
        for col in range(1, col_count + 1):
            ws.cell(row=row, column=col).fill = ALT_ROW_FILL


# ── Section Renderers ────────────────────────────────────────────────────────

def render_key_value(ws, section, start_row):
    """Render a key-value section. Returns next available row."""
    col_count = 2
    apply_section_title(ws, start_row, col_count, section["title"])
    row = start_row + 1

    ws.cell(row=row, column=1, value="字段")
    ws.cell(row=row, column=2, value="值")
    apply_header_style(ws, row, col_count)
    row += 1

    for item in section.get("rows", []):
        ws.cell(row=row, column=1, value=item.get("field", "")).font = LABEL_FONT
        ws.cell(row=row, column=2, value=item.get("value", "")).font = NORMAL_FONT
        ws.cell(row=row, column=2).alignment = WRAP_ALIGNMENT
        apply_cell_border(ws, row, col_count)
        apply_alt_row(ws, row, col_count)
        row += 1

    return row


def render_table(ws, section, start_row):
    """Render a table section. Returns next available row."""
    columns = section.get("columns", [])
    col_count = len(columns)
    apply_section_title(ws, start_row, col_count, section["title"])
    row = start_row + 1

    for ci, col_name in enumerate(columns, 1):
        ws.cell(row=row, column=ci, value=col_name)
    apply_header_style(ws, row, col_count)
    row += 1

    for data_row in section.get("rows", []):
        for ci, val in enumerate(data_row, 1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font = NORMAL_FONT
            cell.alignment = WRAP_ALIGNMENT
        apply_cell_border(ws, row, col_count)
        apply_alt_row(ws, row, col_count)
        row += 1

    return row


def render_narrative(ws, section, start_row):
    """Render a narrative section. Returns next available row."""
    col_count = 1
    apply_section_title(ws, start_row, col_count, section["title"])
    row = start_row + 1

    cell = ws.cell(row=row, column=1, value=section.get("content", ""))
    cell.font = NORMAL_FONT
    cell.alignment = WRAP_ALIGNMENT
    cell.border = THIN_BORDER
    row += 1

    return row


RENDERERS = {
    "key-value": render_key_value,
    "table": render_table,
    "narrative": render_narrative,
}


# ── Main Generation ─────────────────────────────────────────────────────────

def parse_font_spec(spec, default_name="Calibri", default_size=11):
    """Parse a font spec like 'Calibri 11' into name and size."""
    if not spec:
        return default_name, default_size
    parts = spec.strip().rsplit(" ", 1)
    if len(parts) == 2 and parts[1].isdigit():
        return parts[0], int(parts[1])
    return parts[0], default_size


def generate(data, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "DVP"

    # ── Read formatting overrides ────────────────────────────────────────
    fmt = data.get("formatting", {})
    header_color = fmt.get("header_color", "4472C4")
    font_name, font_size = parse_font_spec(fmt.get("font", "Calibri 11"))

    # Override style constants with formatting values
    global HEADER_FILL, HEADER_FONT, SECTION_TITLE_FONT, LABEL_FONT, NORMAL_FONT
    HEADER_FILL = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    HEADER_FONT = Font(name=font_name, size=font_size, bold=True, color="FFFFFF")
    SECTION_TITLE_FONT = Font(name=font_name, size=12, bold=True, color="1F4E79")
    LABEL_FONT = Font(name=font_name, size=font_size, bold=True)
    NORMAL_FONT = Font(name=font_name, size=font_size)

    # ── Meta header ──────────────────────────────────────────────────────
    meta = data.get("meta", {})
    row = 1

    study_name = meta.get("study_name", "未命名研究")
    protocol_no = meta.get("protocol_number", "")
    version = meta.get("version", "")
    date_str = meta.get("date", datetime.now().strftime("%Y-%m-%d"))
    author = meta.get("author", "")

    ws.cell(row=row, column=1, value="DVP - 数据验证计划").font = Font(
        name=font_name, size=16, bold=True, color="1F4E79"
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1

    meta_info = [
        ("研究名称", study_name),
        ("方案编号", protocol_no),
        ("DVP 版本", version),
        ("日期", date_str),
        ("作者", author),
    ]
    for label, value in meta_info:
        ws.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws.cell(row=row, column=2, value=value).font = NORMAL_FONT
        apply_cell_border(ws, row, 2)
        row += 1

    row += 1  # blank row separator

    # ── Sections ─────────────────────────────────────────────────────────
    sections = data.get("sections", [])
    for section in sections:
        section_type = section.get("type", "narrative")
        renderer = RENDERERS.get(section_type, render_narrative)
        row = renderer(ws, section, row)
        row += 1  # blank row between sections

    # ── Auto-width ───────────────────────────────────────────────────────
    max_cols = 1
    for section in sections:
        if section.get("type") == "table":
            max_cols = max(max_cols, len(section.get("columns", [])))
        elif section.get("type") == "key-value":
            max_cols = max(max_cols, 2)
    auto_width(ws, max_cols)

    # ── Print settings ───────────────────────────────────────────────────
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    wb.save(output_path)
    return output_path


# ── CLI Entry Point ──────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python3 generate_xlsx.py <json_file> [--output <output_path>]")
        sys.exit(1)

    json_path = sys.argv[1]

    # Determine output path
    output_path = None
    if "--output" in sys.argv:
        idx = sys.argv.index("--output")
        if idx + 1 < len(sys.argv):
            output_path = sys.argv[idx + 1]

    if not output_path:
        # Try meta.output_path from JSON, then fall back to same-name .xlsx
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                data_preview = json.load(f)
            output_path = data_preview.get("meta", {}).get("output_path", "")
        except Exception:
            output_path = ""
        if not output_path:
            base = os.path.splitext(json_path)[0]
            output_path = base + ".xlsx"

    # Read and validate JSON
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            data = json.load(f)
    except json.JSONDecodeError as e:
        print(f"ERROR: Invalid JSON in {json_path}: {e}")
        sys.exit(1)
    except FileNotFoundError:
        print(f"ERROR: File not found: {json_path}")
        sys.exit(1)

    # Basic validation
    if not isinstance(data, dict):
        print("ERROR: JSON root must be an object")
        sys.exit(1)

    if "sections" not in data:
        print("ERROR: JSON must contain 'sections' array")
        sys.exit(1)

    for i, section in enumerate(data["sections"]):
        if "title" not in section:
            print(f"ERROR: Section {i} missing 'title'")
            sys.exit(1)
        stype = section.get("type", "narrative")
        if stype not in RENDERERS:
            print(f"ERROR: Section {i} has unknown type '{stype}'. Valid: {list(RENDERERS.keys())}")
            sys.exit(1)
        # Validate table row lengths match columns
        if stype == "table" and "columns" in section:
            expected = len(section["columns"])
            for ri, row_data in enumerate(section.get("rows", [])):
                if len(row_data) != expected:
                    print(f"WARNING: Section '{section['title']}' row {ri} has {len(row_data)} values, expected {expected}")

    result = generate(data, output_path)
    section_count = len(data["sections"])
    print(f"OK: Generated {result} with {section_count} sections")


if __name__ == "__main__":
    main()
